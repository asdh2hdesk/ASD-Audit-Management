import base64
from io import BytesIO
from PIL import Image as PILImage
from PIL import ImageOps


from openpyxl import Workbook
from openpyxl.drawing.image import Image

from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from odoo import api, fields, models
import io

class IATFAuditChecklist(models.Model):
    _inherit = 'mgmtsystem.audit.auditor'
    company_id = fields.Many2one('res.company', string='Company', required=True, default=lambda self: self.env.company)
    generate_xls_file = fields.Binary(string="Generated File")

    def action_generate_audit_excel_report(self):
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "IMS Audit Checksheet"

        # Add company logo if it exists
        if self.env.user.company_id.logo:
            try:
                image_data = base64.b64decode(self.env.user.company_id.logo)
                logo_image = self._process_company_logo(image_data)
                ws.add_image(logo_image, 'H3')
            except Exception as e:
                raise ValueError("Error adding company logo to the worksheet.") from e

        # Define styles
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        font_header = Font(name='Arial', size=12, bold=True)



        # Set column widths
        self._set_column_widths(ws)

        # Write headers and metadata
        self._write_headers(ws, align_center, font_header, border)

        # Write data rows
        self._fill_audit_metadata(ws)
        self._fill_audit_checklist(ws, border, align_center)

        # Save the workbook to a binary stream
        wb.save(output)
        output.seek(0)

        # Set the generated file
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        # Return download URL
        return {
            'type': 'ir.actions.act_url',
            'target': 'self',
            'url': f'/web/content?model=mgmtsystem.audit.auditor&id={self.id}&filename=IMS_Audit_Checksheet.xlsx&field=generate_xls_file&download=true'
        }

    def _process_company_logo(self, image_data):
        """Process and resize the company logo for insertion into the Excel sheet."""
        max_width, max_height = 200, 100
        image = PILImage.open(BytesIO(image_data))
        width, height = image.size
        aspect_ratio = width / height

        # Resize the image while maintaining the aspect ratio
        if width > max_width:
            width = max_width
            height = int(width / aspect_ratio)
        if height > max_height:
            height = max_height
            width = int(height * aspect_ratio)

        resized_image = image.resize((width, height), PILImage.LANCZOS)

        # Add padding to the image
        padding_top = 5
        padding_left = 5
        resized_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill=(255, 255, 255, 0))

        # Convert the image to a binary stream
        img_bytes = BytesIO()
        resized_image.save(img_bytes, format='PNG')
        img_bytes.seek(0)

        return Image(img_bytes)

    def _set_column_widths(self, ws):
        """Set the widths of the columns in the worksheet."""
        column_widths = [15, 25, 20, 20, 20, 25, 20, 25, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _write_headers(self, ws, align_center, font_header, border):
        """Write the headers and metadata to the worksheet."""
        # Main title
        ws.merge_cells('A1:I2')
        ws['A1'] = 'IMS  INTERNAL  AUDIT  CHECKSHEET'
        ws['A1'].font = Font(color="FFFFFF", bold=True,size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

        # Logo placeholder
        ws.merge_cells('H3:I7')
        ws['H3'] = 'Your logo'
        ws['H3'].alignment = align_center
        ws['H3'].border = border

        # # Metadata rows
        # metadata = [
        #     ('Company Name', None, None),
        #     ('Department', None, None),
        #     ('Last TUV Audit Observations/NC', None,  None),
        #     ('Last Customer Audit Observations/NC', None, None),
        #     ('Last Internal Audit Observations/NC', None, None),
        # ]
        # Define fill color
        fill_light_blue = PatternFill(start_color='C2DCEC', end_color='C2DCEC', fill_type='solid')
        font_bold_blue = Font(name='Arial', size=11, bold=True, color="1C73C4")

        # Define the data dictionary
        data = {
            'E3': 'Target Date',
            'E4': 'Audit',
            'E5': 'Auditor',
            'E6': 'Auditee',
            'E7': 'Audit Reference',
            'A3': 'Company Name',
            'A4': 'Department',
            'A5': 'Last TUV Audit Observations/NC',
            'A6': 'Last Customer Audit Observations/NC',
            'A7': 'Last Internal Audit Observations/NC',
        }

        # Write the data into the worksheet
        for cell, value in data.items():
            ws[cell] = value
            ws[cell].fill = fill_light_blue
            ws[cell].font = font_bold_blue  # Apply blue font

        # Define styles
        thin = Side(border_style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_bold = Font(name='Aerial', size=11, bold=True)
        align_center = Alignment(vertical='center', horizontal='center', wrapText=True)
        align_left = Alignment(vertical='center', horizontal='left', wrapText=True)

        # Write the data into the worksheet and apply styles
        for cell, value in data.items():
            ws[cell] = value  # Write value to the specified cell
            ws[cell].border = border  # Apply border to the cell
            if cell in ['A5', 'A6', 'A7']:  # For specific cells requiring left alignment and bold text
                ws[cell].alignment = align_left




        # Merge cells dynamically based on row alignment
        for cell in data.keys():
            column, row = cell[:1], int(cell[1:])  # Extract column and row
            if column == 'A':  # Merge for 'A' column labels
                ws.merge_cells(f'A{row}:B{row}')
            elif column == 'C':  # Merge for 'C' column labels
                ws.merge_cells(f'C{row}:D{row}')
            elif column == 'F':  # Merge for 'F' column labels
                ws.merge_cells(f'F{row}:G{row}')
            # Apply styles
            ws[cell].alignment = align_center
            ws[cell].font = font_header
            ws[cell].border = border
            for cell, value in data.items():
                ws[cell] = value  # Write value to the specified cell
                ws[cell].font = font_bold  # Make text bold
                ws[cell].alignment = align_left if cell in ['A5', 'A6', 'A7'] else align_center

            # Merge cells dynamically for columns C-D and F-G
            for row in range(3, 8):  # Adjust the range as needed
                ws.merge_cells(f'C{row}:D{row}')  # Merge columns C-D for each row
                ws.merge_cells(f'F{row}:G{row}')  # Merge columns F-G for each row


            # Apply borders to all cells in the sheet
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

            ws.merge_cells('A8:I8')
            ws['A8'].fill = fill_light_blue
            ws['A8'].alignment = align_center
            ws['A8'].border = border
            ws['A8'].font = font_bold

            # Set labels
            # ws[f'A{row}'] = left_label
            # ws[f'E{row}'] = right_label


            # Apply styling
            # for cell in [ws[f'A{row}'], ws[f'E{row}'], ws[f'H{row}']]:
            #     cell.alignment = align_center
            #     cell.border = border

        # Add table headers
        checklist_headers = [
            'Clause', 'Checkpoints', 'Department', 'Procedure', 'Format Number', 
            'Observations/Findings', 'Criteria', 'Objective Evidence', 'Status'
        ]
        header_row = len(data) - 1 # Calculate the header row number
        for col, header in enumerate(checklist_headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = align_center
            cell.border = border

    def _fill_audit_checklist(self, ws, border, align_center, row_idx=None):
        """Fill the audit checklist data rows in the worksheet."""
        cur_row = 10  # Start from row 10 (after headers in row 9)
        for line in self.filtered_line_ids:
            # Populate data cells
            ws[f'A{cur_row}'] = line.clause if line.clause else ''
            ws[f'B{cur_row}'] = line.checkpoints if line.checkpoints else ''
            ws[f'C{cur_row}'] = line.department_id.name if line.department_id else ''
            ws[f'D{cur_row}'] = line.observations if line.observations else ''
            ws[f'E{cur_row}'] = line.procedure_id.name if line.procedure_id else ''
            ws[f'F{cur_row}'] = line.format_number if line.format_number else ''
            ws[f'G{cur_row}'] = line.criteria if line.criteria else ''
            ws[f'H{cur_row}'] = line.objective_evidence if line.objective_evidence else ''
            ws[f'I{cur_row}'] = line.status if line.status else ''

            if line.evidence_images_ids:
                for attachment in line.evidence_images_ids:
                    if attachment.mimetype and "image" in attachment.mimetype:
                        # try:
                            # Decode the base64 image data
                            image_data = base64.b64decode(attachment.datas)

                            # Resize the image
                            max_width = 140 # Maximum width in pixels
                            max_height = 20# Maximum height in pixels

                            # Open the image using PIL
                            image = PILImage.open(io.BytesIO(image_data))
                            width, height = image.size
                            aspect_ratio = width / height

                            # Resize while maintaining aspect ratio
                            if width > max_width:
                                width = max_width
                                height = int(width / aspect_ratio)

                            if height > max_height:
                                height = max_height
                                width = int(height * aspect_ratio)

                            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)

                            # Convert resized image to BytesIO for OpenPyXL
                            img_bytes = io.BytesIO()
                            resized_image.save(img_bytes, format='PNG')
                            img_bytes.seek(0)

                            # Add the image to the worksheet
                            cell_location = f'H{cur_row}'
                            img = Image(img_bytes)
                            ws.add_image(img, cell_location)

                            # Adjust row height to fit the image
                            ws.row_dimensions[cur_row].height = 80  # Adjust as needed

                        # except Exception as e:
                        #     _logger.error(f"Error processing image {attachment.name}: {e}")

            # Apply borders and alignment to all cells in the row
            for col in range(1, 10):
                cell = ws.cell(row=cur_row, column=col)
                cell.alignment = align_center
                cell.border = border
                # Example for multiple cells
                cell.alignment = Alignment(wrap_text=True)

            cur_row += 1


    def _fill_audit_metadata(self, ws):
        """Fill the metadata fields in the worksheet."""
        # Populate header information
        ws['C3'] = self.company_id.name if self.company_id else ''  # Company Name
        ws['C4'] = self.department_id.name if self.department_id else ''  # Department
        ws[
            'C5'] = self.last_tuv_audit_obs if self.last_tuv_audit_obs else ''  # Last TUV Audit Observations/NC
        ws[
            'C6'] = self.last_customer_audit_obs if self.last_customer_audit_obs else ''  # Last Customer Audit Observations/NC
        ws[
            'C7'] = self.last_internal_ims_audit_obs if self.last_internal_ims_audit_obs else ''  # Last Internal Audit Observations/NC

        ws['F3'] = self.target_date.strftime('%Y-%m-%d') if self.target_date else 'N/A'     # Date
        ws['F4'] = self.audit_id.name if self.audit_id else ''      # Shift
        ws['F5'] = self.name if self.name else ''  # Auditor
        ws['F6'] = ','.join([auditee.name for auditee in self.selected_user_ids])  # Auditee
        ws['F7'] = self.audit_ref if self.audit_ref else ''
