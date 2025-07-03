import base64
import logging
import io
from io import BytesIO
from PIL import ImageOps, Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from odoo import api, fields, models


class RevisionChecksheet(models.Model):
    _inherit = 'audit.revision.data'

    generate_xls_file = fields.Binary(string="Generated File", readonly=True)
    xls_filename = fields.Char(string="Filename", default="Audit_Revision_Report.xlsx")

    def action_generate_revision_report(self):
        for rec in self:
            wb = Workbook()

            # Remove default sheet
            if wb.active.title == 'Sheet':
                wb.remove(wb.active)

            # Create Audit Revision Data Sheet
            ws = wb.create_sheet("Audit Revision Data")
            self._style_and_fill_sheet(ws, rec)


            # Save the workbook
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Encode the file in base64
            rec.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

            return {
                'type': 'ir.actions.act_url',
                'target': 'self',
                'url': f'/web/content?model=audit.revision.data&id={rec.id}&filename=Audit_Revision_Report.xlsx&field=generate_xls_file&download=true'
            }



    def _style_and_fill_sheet(self, ws, rec):
        """ Apply styles and fill the Excel sheet with revision data """
        styles = self._get_common_styles()

        # Set column widths
        for col in 'AB':
            ws.column_dimensions[col].width = 40
            ws.column_dimensions[col].height=20

        # Title
        ws.merge_cells('A1:B2')
        ws['A1'] = "Audit Revision Data"
        ws['A1'].font = styles['header_font']
        ws['A1'].fill = styles['header_fill']
        ws['A1'].alignment = styles['align_center']

        # Data Fields
        headers = [
            ('A3', 'Audit:', 'B3', rec.audit_id.name if rec.audit_id else 'N/A'),
            ('A4', 'Reference:', 'B4', rec.reference if rec.reference else 'N/A'),
            ('A5', 'Revision Data:', 'B5', rec.revision_data if rec.revision_data else 'N/A'),
            ('A6', 'Created on:', 'B6', rec.create_date.strftime('%Y-%m-%d') if rec.create_date else "N/A"),
            ('A7', 'Created By:', 'B7', rec.create_uid.name if rec.create_uid else 'N/A')
        ]

        for label_ref, label, value_ref, value in headers:
            ws[label_ref] = label
            ws[label_ref].font = styles['header_font']
            ws[label_ref].fill = styles['header_fill']
            ws[label_ref].border = styles['border']
            ws[label_ref].alignment = styles['align_center']

            ws[value_ref] = value
            ws[value_ref].font = styles['normal_font']
            ws[value_ref].border = styles['border']
            ws[value_ref].alignment = styles['align_center']

    def _get_common_styles(self):
        """ Common styles for formatting the Excel sheets """
        return {
            'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'table_header_fill': PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'header_font': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'normal_font': Font(name='Arial', size=13),
            'align_left': Alignment(horizontal='left', vertical='center'),
            'align_center': Alignment(horizontal='center', vertical='center')
        }
