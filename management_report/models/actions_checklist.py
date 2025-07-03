import base64
import logging
from io import BytesIO
from PIL import Image as PILImage
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from odoo import api, fields, models
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup
import openpyxl


class ActionsChecksheet(models.Model):
    _inherit = 'mgmtsystem.action'
    generate_xls_file = fields.Binary(string="Generated File")

    def action_generate_audit_report(self):
        for rec in self:
            # Create a new workbook
            wb = openpyxl.Workbook()



            # Remove default sheet if not used
            if wb.active.title == 'Sheet':
                wb.remove(wb.active)

            # Add Description sheet
            ws_description = wb.create_sheet("Description")
            self.get_resized_img(self.env.user.company_id.logo, 'F3',ws_description)
            self._create_description_sheet(ws_description, rec)

            # Add Related Nonconformities sheet
            ws_nonconformities = wb.create_sheet("Related Nonconformities")
            self.get_resized_img(self.env.user.company_id.logo, 'F3',ws_nonconformities)
            self._create_related_nonconformities_sheet(ws_nonconformities, rec)


            # Save the workbook to a BytesIO object
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Encode the file in base64 and attach it to the record
            rec.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

            return {
                'type': 'ir.actions.act_url',
                'target': 'self',
                'url': f'/web/content?model=mgmtsystem.action&id={rec.id}&filename=Action_Report.xlsx&field=generate_xls_file&download=true'
            }


    def parse_html_to_text_list(self, html_content):
        """
        Parse HTML content and return a list of text elements.

        :param html_content: str, HTML content to parse
        :return: list of str, extracted text elements
        """
        # Create a BeautifulSoup object
        soup = BeautifulSoup(html_content, 'html.parser')

        # Extract text from the soup object
        text_elements = soup.find_all(text=True)

        # Filter out unwanted elements (like script, style, etc.)
        filtered_text = [text.strip() for text in text_elements if
                         text.strip() and text.parent.name not in ['script', 'style']]

        return filtered_text

    def _create_description_sheet(self, ws, rec):
        """Create the Description sheet with the specified format."""
        """Create the Description sheet with the specified format."""
        styles = self._get_common_styles()

        # Define the background color for the header
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

        # Write the Action Title
        ws.merge_cells('A1:H2')
        ws['A1'] = "Action Report"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = header_fill

        # Define the structure
        fields = [

            ("A3", "Response Type", "C3", rec.type_action or "N/A"),
            ("D3", "Sequence", "E3", rec.sequence or "N/A"),
            ("A4", "Responsible", "C4", rec.user_id.name or "N/A"),
            ("D4", "System", "E4", rec.system_id.name or "N/A"),
            ("A5", "Deadline", "C5", rec.date_deadline or "N/A"),
            ("D6", "Reference", "E6", rec.reference or "N/A"),
            ("D5", "Created On", "E5", rec.create_date.strftime('%Y-%m-%d') if rec.create_date else "N/A"),
            ("A6", "Tags", "C6", ', '.join(tag.name for tag in rec.tag_ids) if rec.tag_ids else "N/A"),
        ]

        # Populate the table
        for label_cell, label, value_cell, value in fields:
            ws.merge_cells('A3:B3'),
            ws.merge_cells('A4:B4'),
            ws.merge_cells('A5:B5'),
            ws.merge_cells('A6:B6'),
            ws[label_cell] = label
            ws[label_cell].font = styles['header_font']
            ws[label_cell].fill = header_fill
            ws[label_cell].alignment = styles.get('align_center', Alignment(horizontal='center', vertical='center'))

            ws[label_cell].border = styles['border']

            ws[value_cell] = value
            ws[value_cell].font = styles['normal_font']
            ws[value_cell].alignment = styles['align_left']
            ws[value_cell].border = styles['border']

        # Add placeholder for Description
        ws.merge_cells('A7:H7')
        ws['A7'] = "Description"
        ws['A7'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws['A7'].fill = header_fill
        ws['A7'].alignment = styles.get('align_center', Alignment(horizontal='center', vertical='center'))
        ws['A7'].border = styles['border']

        # Adjust row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[7].height = 25

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

        # Apply wrap text and alignment
        for row in ws.iter_rows(min_row=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Add description data below the "Description" header
        row = 8  # Start adding description data from row 8
        if rec.description:
            # Parse HTML content into a list of text elements
            description_parsed = self.parse_html_to_text_list(rec.description)

            if description_parsed:
                # Split the parsed content into key-value pairs
                description_dict = {}
                current_key = None

                for line in description_parsed:
                    if ":" in line:  # Check if the line contains a key-value pair
                        key, value = line.split(":", 1)  # Split only on the first occurrence of ":"
                        description_dict[key.strip()] = value.strip()
                        current_key = key.strip()
                    elif current_key:  # If the line is a continuation of the previous value
                        description_dict[current_key] += " " + line.strip()

                # Insert parsed description data in multiple rows
                for key, value in description_dict.items():
                    # Add key in column A
                    ws[f'A{row}'] = key
                    ws[f'A{row}'].alignment = styles['align_left']
                    ws[f'A{row}'].font = styles['normal_font']
                    ws[f'A{row}'].border = styles['border']

                    # Add value in column B
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].alignment = styles['align_left']
                    ws[f'B{row}'].font = styles['normal_font']
                    ws[f'B{row}'].border = styles['border']
                    ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    ws.merge_cells(f'B{row}:H{row}')
                    row += 1
            else:
                # If no parsed content, add a placeholder
                ws[f'A{row}'] = "No description available"
                ws[f'A{row}'].alignment = styles['align_left']
                ws[f'A{row}'].font = styles['normal_font']
                ws[f'A{row}'].border = styles['border']

                ws.merge_cells(f'B{row}:H{row}')
                ws[f'B{row}'] = "No description available"
                ws[f'B{row}'].alignment = styles['align_left']
                ws[f'B{row}'].font = styles['normal_font']
                ws[f'B{row}'].border = styles['border']
                ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            # If no description, add a placeholder
            ws[f'A{row}'] = "No description available"
            ws[f'A{row}'].alignment = styles['align_left']
            ws[f'A{row}'].font = styles['normal_font']
            ws[f'A{row}'].border = styles['border']

            ws.merge_cells(f'B{row}:H{row}')
            ws[f'B{row}'] = "No description available"
            ws[f'B{row}'].alignment = styles['align_left']
            ws[f'B{row}'].font = styles['normal_font']
            ws[f'B{row}'].border = styles['border']
            ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Adjust row heights
        ws.row_dimensions[1].height = 30  # Action Report title
        ws.row_dimensions[7].height = 25  # Description header
        # for r in range(8, row + 1):  # Adjust rows for description data
        #     ws.row_dimensions[r].height = 20

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

        # Apply wrap text and alignment
        # for row in ws.iter_rows():
        #     for cell in row:
        #         cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _create_related_nonconformities_sheet(self, ws, rec):
        """Create the Related Nonconformities sheet with the specified format."""
        styles = self._get_common_styles()
        ws.title = "Related Nonconformities"

        # Define the background color for the header
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

        # Write the Action Title
        ws.merge_cells('A1:H2')
        ws['A1'] = "Action Report"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = header_fill

        # Define the structure
        fields = [
            ("A3", "Response Type", "C3", rec.type_action or "N/A"),
            ("D3", "Sequence", "E3", rec.sequence or "N/A"),
            ("A4", "Responsible", "C4", rec.user_id.name or "N/A"),
            ("D4", "System", "E4", rec.system_id.name or "N/A"),
            ("A5", "Deadline", "C5", rec.date_deadline or "N/A"),
            ("D6", "Reference", "E6", rec.reference or "N/A"),
            ("D5", "Created On", "E5", rec.create_date.strftime('%Y-%m-%d') if rec.create_date else "N/A"),
            ("A6", "Tags", "C6", ', '.join(tag.name for tag in rec.tag_ids) if rec.tag_ids else "N/A"),
        ]

        # Populate the table
        for label_cell, label, value_cell, value in fields:
            ws.merge_cells('A3:B3'),
            ws.merge_cells('A4:B4'),
            ws.merge_cells('A5:B5'),
            ws.merge_cells('A6:B6'),
            ws[label_cell] = label
            ws[label_cell].font = styles['header_font']
            ws[label_cell].fill = header_fill
            ws[label_cell].alignment = styles.get('align_center', Alignment(horizontal='center', vertical='center'))



            ws[label_cell].border = styles['border']

            ws[value_cell] = value
            ws[value_cell].font = styles['normal_font']
            ws[value_cell].alignment = styles['align_left']
            ws[value_cell].border = styles['border']

        # Add headers for nonconformities
        headers = ["Reference", "Name", "Created On", "Filled In By", "Responsible", "Manager", "System", "Stage"]
        ws.append(headers)
        ws.row_dimensions[8].height = 30
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws['A7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['B7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['C7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['D7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['E7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['F7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['G7'].font = Font(name='Arial', size=14, bold=True, color='060115')
        ws['H7'].font = Font(name='Arial', size=14, bold=True, color='060115')


        # Apply header formatting
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=7, column=col)
            cell.font = styles['header_font']
            cell.fill = header_fill
            # cell.alignment = styles['align_center']
            cell.border = styles['border']
            cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')

        # Write data
        for nc in rec.nonconformity_ids:
            ws.append([
                nc.ref or "N/A",
                nc.name or "N/A",
                nc.create_date.strftime('%Y-%m-%d') if nc.create_date else "N/A",
                nc.user_id_by or "N/A",
                nc.responsible_user_id.name or "N/A",
                nc.manager_user_id.name or "N/A",
                nc.system_id.name or "N/A",
                nc.stage_id or "N/A"
            ])

        # Add description date in A8
        ws['A8'] = rec.date_deadline.strftime('%Y-%m-%d') if rec.date_deadline else "N/A"
        ws['A8'].font = Font(name='Arial', size=11)
        ws['A8'].alignment = Alignment(wrap_text=True, vertical='top')
        ws['A8'].border = styles['border']

        # Apply wrap text and alignment
        # for row in ws.iter_rows():
        #     for cell in row:
        #         cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _get_common_styles(self):
        """Define common styles for the report."""
        return {
            'header_font': Font(name='Arial', size=11, bold=True),
            'normal_font': Font(name='Arial', size=10),
            'align_left': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        }
    # def _process_company_logo(self, image_data):
    #     """Process and resize the company logo for insertion into the Excel sheet."""
    #     max_width, max_height = 200, 100
    #     image = PILImage.open(BytesIO(image_data))
    #     width, height = image.size
    #     aspect_ratio = width / height

    def get_resized_img(self, image, cell, ws, max_width=350, max_height=80, padding_top=10, padding_left=10):

        # if not self.env.user.company_id.logo:
        #     return None  # No logo available

        image_data = base64.b64decode(image)
        image = PILImage.open(io.BytesIO(image_data))

        # Get original dimensions and aspect ratio
        width, height = image.size
        aspect_ratio = width / height

        # Resize while maintaining aspect ratio
        if width > max_width:
            width = max_width
            height = int(width / aspect_ratio)

        if height > max_height:
            height = max_height
            width = int(height * aspect_ratio)

        # Resize image
        resized_image = image.resize((width, height), PILImage.LANCZOS)

        # Add padding (transparent background)
        padded_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill=(255, 255, 255, 0))

        # Save to bytes
        img_bytes = io.BytesIO()
        padded_image.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        logo_image = Image(img_bytes)
        ws.add_image(logo_image, cell)

    def add_logo_to_worksheet(self, ws, cell="F3"):

        logo_image = self.get_resized_logo()
        if logo_image:
            ws.add_image(logo_image, cell)

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            # Apply the border to the cells where the logo is placed
            ws[cell].border = border

    def _process_image(self, image_data, max_size=(100, 90)):
        # try:
            img = PILImage.open(BytesIO(image_data))
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            temp_bytes = BytesIO()
            img.save(temp_bytes, format='PNG')
            return Image(temp_bytes)
        # except Exception as e:
        #     _logger.error("Image processing failed: %s", str(e))
        #     return None




