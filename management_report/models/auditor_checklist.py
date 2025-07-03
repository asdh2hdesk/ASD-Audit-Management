import base64
import logging
from io import BytesIO
from PIL import Image as PILImage
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from odoo import api, fields, models
from PIL import ImageOps
from PIL import Image as PILImage
from bs4 import BeautifulSoup

_logger = logging.getLogger(__name__)


class AuditChecklist(models.Model):
    _inherit = 'mgmtsystem.audit'

    # Define all required relational fields
    # strong_points_ids = fields.One2many("strong.point.line", "audit_id", "Strong Points")
    #auditor_user_ids = fields.One2many("auditor.line", "audit_id", "Auditors")
    # line_ids = fields.One2many("verification.line", "audit_id", "Verification List")
    # nonconformity_ids = fields.One2many("nonconformity.line", "audit_id", "Nonconformities")
    # to_improve_points = fields.One2many("improve.point.line", "audit_id", "To Improve Points")
    # imp_opp_ids = fields.One2many("improvement.opp.line", "audit_id", "Improvement Opportunities")
    # audit_id = fields.Many2one('mgmtsystem.audit', string='Audit')
    generate_xls_file = fields.Binary(string="Generated File")

    def action_generate_audit_report(self):
        for rec in self:
            print(rec.auditee_user_ids)
        output = BytesIO()
        wb = Workbook()

        # Remove default sheet if not used
        if wb.active.title == 'Sheet':
            wb.remove(wb.active)

        sheets = {
            'Auditors Report': self._create_auditors_report,
            'Auditees Report': self._create_auditees_report,
            'Verification List': self._create_verification_list,
            'Strong Points': self._create_strong_points,
            'To Improve Points': self._create_to_improve_points,
            'Improvement Opportunities': self._create_improvement_opportunities,
            'Nonconformities': self._create_nonconformities_sheet,
        }

        for sheet_name, create_func in sheets.items():
            try:
                ws = wb.create_sheet(sheet_name)
                create_func(ws)
                self.get_resized_img(self.env.user.company_id.logo, 'E1', ws)
            except Exception as e:
                _logger.error("Error creating sheet %s: %s", sheet_name, str(e))

        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        return {
            'type': 'ir.actions.act_url',
            'target': 'self',
            'url': f'/web/content?model=mgmtsystem.audit&id={self.id}&filename=Audit_Report.xlsx&field=generate_xls_file&download=true'
        }

    def _get_common_styles(self):
        return {
            'header_fill': PatternFill(start_color='C2DCEC', end_color='C2DCEC', fill_type='solid'),
            'table_header_fill': PatternFill(start_color="003366", end_color="003366", fill_type="solid"),

            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'header_font': Font(name='Arial', size=11, bold=True,color='000000'),
            'table_header_font': Font(name='Arial', size=11,bold=True, color='FFFFFF'),
            'normal_font': Font(name='Arial', size=10),
            'align_left': Alignment(horizontal='left', vertical='center'),
            'align_center': Alignment(horizontal='center', vertical='center')
        }

    def _set_common_header(self, ws):
        styles = self._get_common_styles()
        # Set column widths
        for col in 'ABCD':
            ws.column_dimensions[col].width = 25

        headers = [
            ('A1', 'Name:', 'B1', self.name or 'N/A'),
            ('A2', 'Reference:', 'B2', self.reference or 'N/A'),
            ('A3', 'Audit Manager:', 'B3', self.user_id.name if self.user_id else 'N/A'),
            ('A4', 'Planned Date:', 'B4', self.date.strftime('%Y-%m-%d') if self.date else 'N/A'),
            ('C1', 'Company:', 'D1', self.company_id.name if self.company_id else 'N/A'),
            ('C2', 'City:', 'D2', self.company_state if self.company_state else 'N/A'),
            ('C3', 'Plant Code:', 'D3', self.plant_name or 'N/A'),
            ('C4', 'System:', 'D4', self.system_id.name if self.system_id else 'N/A')
        ]
        # last_col_index = len(headers)  # Get last column index
        # last_col_letter = get_column_letter(last_col_index)

        for label_ref, label, value_ref, value in headers:
            label_cell = ws[label_ref]
            label_cell.value = label
            label_cell.font = styles['header_font']
            label_cell.fill = styles['header_fill']
            label_cell.border = styles['border']
            label_cell.alignment = styles['align_left']

            value_cell = ws[value_ref]
            value_cell.value = value
            value_cell.font = styles['normal_font']
            value_cell.border = styles['border']
            value_cell.alignment = styles['align_left']

        # # Add company logo
        # if self.env.company.logo:
        #     max_width = 500  # Set your desired maximum width
        #     max_height = 100  # Set your desired maximum height
        #     image_data = base64.b64decode(self.env.user.company_id.logo)
        #
        #     # Open the image using PIL
        #     image = PILImage.open(io.BytesIO(image_data))
        #     width, height = image.size
        #     aspect_ratio = width / height
        #
        #     if width > max_width:
        #         width = max_width
        #         height = int(width / aspect_ratio)
        #
        #     if height > max_height:
        #         height = max_height
        #         width = int(height * aspect_ratio)
        #
        #     # Resize the image using PIL
        #     # Add space on the top and left side of the image
        #     padding_top = 10  # Adjust as needed
        #     padding_left = 10  # Adjust as needed
        #
        #     resized_image = image.resize((width, height), PILImage.LANCZOS)
        #     ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
        #     img_bytes = io.BytesIO()
        #     resized_image.save(img_bytes, format='PNG')
        #     img_bytes.seek(0)
        #     logo_image = Image(img_bytes)
        #
        #     # Decode and process the company logo
        #     logo_img = self._process_company_logo(base64.b64decode(self.env.company.logo))
        #     if logo_img:
        #         ws.add_image(logo_img, 'D1')  # Add image to cell D1
        #         ws.row_dimensions[1].height = 120  # Adjust row height for better display
        #         ws.column_dimensions['D'].width = 20
        #         ws.merge_cells('D1:D4'),
        #         # Adjust column width if necessary


    def _create_sheet_structure(self, ws, columns):
        styles = self._get_common_styles()
        self._set_common_header(ws)

        # Set column widths and headers
        for col, (header, width) in columns.items():
            ws.column_dimensions[col].width = width
            cell = ws.cell(row=5, column=ord(col) - ord('A') + 1, value=header)
            cell.font = styles['table_header_font']
            cell.fill = styles['table_header_fill']
            cell.border = styles['border']
            cell.alignment = styles['align_center']

            cell.alignment = Alignment(wrap_text=True)

    # ----- Auditors Report -----
    def _create_auditors_report(self, ws):
        columns = {
            'A': ('Auditor Name', 20),
            'B': ('Target Date', 20),
            'C': ('Auditor Department', 30),
            'D': ('Department to Audit', 40),
            'E': ('Certification Date', 20),
            'F': ('Certification Expiry Date', 40),
            'G': ('Status', 20)
        }
        self._create_sheet_structure(ws, columns)
        self._fill_auditors_report(ws)

    def _fill_auditors_report(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for auditor in self.auditors_user_ids:
                ws[f'A{cur_row}'] = auditor.user_id.name if auditor.user_id else ''
                ws[f'B{cur_row}'] = auditor.target_date.strftime('%Y-%m-%d') if auditor.target_date else ''
                ws[f'C{cur_row}'] = auditor.department_id.name if auditor.department_id else ''
                ws[f'D{cur_row}'] = ', '.join(auditor.department_ids.mapped('name')) if auditor.department_ids else ''
                ws[f'E{cur_row}'] = auditor.certified_date.strftime('%Y-%m-%d') if auditor.certified_date else ''
                ws[f'F{cur_row}'] = auditor.expiry_date.strftime('%Y-%m-%d') if auditor.expiry_date else ''
                ws[f'G{cur_row}'] = auditor.state or ''

                # last_col = len(columns)  # Get last used column index
                # last_col_letter = chr(64 + last_col)  # Convert number to column letter (A, B, C...)
                #
                # # Merge header row dynamically from A1 to the last column
                # merge_range = f"A1:{last_col_letter}1"
                # ws.merge_cells(merge_range)

                # Apply styling
                for col in 'ABCDEFG':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                cur_row += 1
        except Exception as e:
            _logger.error("Error filling auditors report: %s", str(e))

    # ----- Auditees Report -----
    def _create_auditees_report(self, ws):
        columns = {
            'A': ('Auditee Name', 30),
            'C': ('Auditee Department', 30)
        }

        self._create_sheet_structure(ws, columns)
        self._fill_auditees_report(ws)
        ws.merge_cells('A5:B5')
        ws.merge_cells('C5:D5')

        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        # Auto-adjust column width based on header text length
       # for col, (header, default_width) in columns.items():
       #      ws.column_dimensio ns[col].width = max(default_width, len(header) + 5)

    def _fill_auditees_report(self, ws):

        styles = self._get_common_styles()
        cur_row = 6

        try:
            for auditee in self.auditees_user_ids:  # Ensure auditee_ids field exists
                ws.merge_cells(f'A{cur_row}:B{cur_row}')
                ws.merge_cells(f'C{cur_row}:D{cur_row}')
                ws[f'A{cur_row}'] = auditee.user_id.name if auditee.user_id else ''
                ws[f'A{cur_row}'].border = styles['border']
                ws[f'A{cur_row}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                ws[f'C{cur_row}'] = auditee.department_id.name if auditee.department_id else ''
                ws[f'C{cur_row}'].border = styles['border']

                ws[f'C{cur_row}'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        except Exception as e:
            _logger.error("Error filling auditees report: %s", str(e))

            for col in 'ABCDEF':
                  cell = ws[f'{col}{cur_row}']
                  cell.border = styles['border']
                  cell.alignment = styles['align_center']
            cur_row += 1


    # ----- To Improve Points -----
    def _create_to_improve_points(self, ws):
        columns = {
            'A': ('Clause', 20),
            'B': ('Checkpoints', 40),
            'D': ('Observations/Findings', 40),
            'F': ('Status', 20)

        }
        self._create_sheet_structure(ws, columns)
        self._fill_to_improve_points(ws)
        ws.merge_cells('B5:C5')
        ws.merge_cells('D5:E5')
        ws['B5'].alignment = Alignment( horizontal='center', vertical='center')
        ws['D5'].alignment = Alignment( horizontal='center', vertical='center')


    def _fill_to_improve_points(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for point in self.to_improve_points_ids:
                ws.merge_cells(f'B{cur_row}:C{cur_row}')
                ws.merge_cells(f'D{cur_row}:E{cur_row}')
                ws[f'A{cur_row}'] = point.clause or ''
                ws[f'B{cur_row}'] = point.checkpoints if point.checkpoints else ''
                ws[f'D{cur_row}'] = point.observations if point.observations else ''
                ws[f'E{cur_row}'] = point.status if point.status else ''
                # ws[f'A{cur_row}'].alignment = Alignment( horizontal='center', vertical='center')
                # ws[f'D{cur_row}'].alignment = Alignment( horizontal='center', vertical='center')




                for col in 'ABCDEFG':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cur_row += 1
        except Exception as e:
            _logger.error("Error filling to improve points: %s", str(e))

    # ----- Improvement Opportunities -----
    def _create_improvement_opportunities(self, ws):
        columns = {
            'A': ('Reference', 30),
            'B': ('Subject', 40),
            'C': ('Response Type', 40),
            'D': ('Responsible', 30),
            'E': ('Deadline', 30),
            'F': ('Stage', 20)
        }
        self._create_sheet_structure(ws, columns)
        self._fill_improvement_opportunities(ws)

    def _fill_improvement_opportunities(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for opp in self.imp_opp_ids:
                ws[f'A{cur_row}'] = opp.reference or ''
                ws[f'B{cur_row}'] = opp.name or ''
                ws[f'C{cur_row}'] = opp.type_action if opp.type_action else ''
                ws[f'D{cur_row}'] = opp.user_id.name if opp.user_id else ''
                ws[f'E{cur_row}'] = opp.date_deadline.strftime('%Y-%m-%d') if opp.date_deadline else ''
                ws[f'F{cur_row}'] = opp.stage_id.name if opp.stage_id else ''

                for col in 'ABCDEF':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.border = styles['border']

                cur_row += 1
        except Exception as e:
            _logger.error("Error filling improvement opportunities: %s", str(e))
    # ----- Verification List -----
    def _create_verification_list(self, ws):
        columns = {
            'A': ('Clause', 20),
            'B': ('Checkpoints', 40),
            'C': ('Department', 40),
            'D': ('Procedure', 30),
            'E': ('Format Number', 40),
            'F': ('Observations/Findings', 50),
            'G': ('Criteria', 20),
            'H': ('Objective Evidence', 20),
            'I': ('Evidence Images', 20),
            'J': ('Status', 20)
        }
        self._create_sheet_structure(ws, columns)
        self._fill_verification_list(ws)

    def _fill_verification_list(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for line in self.line_ids:
                ws[f'A{cur_row}'] = line.clause if line.clause else ''
                ws[f'B{cur_row}'] = line.checkpoints if line.checkpoints else ''
                ws[f'C{cur_row}'] = line.department_id.name if line.department_id else ''
                ws[f'D{cur_row}'] = line.observations if line.observations else ''
                ws[f'E{cur_row}'] = line.procedure_id.name if line.procedure_id else ''
                ws[f'F{cur_row}'] = line.format_number if line.format_number else ''
                ws[f'G{cur_row}'] = line.criteria if line.criteria else ''
                ws[f'H{cur_row}'] = line.objective_evidence if line.objective_evidence else ''
               # ws[f'I{cur_row}'] = line.evidence_images if line.evidence_images else ''
                ws[f'J{cur_row}'] = line.status if line.status else ''
                # Handle image
                if line.evidence_images_ids:
                    for attachment in line.evidence_images_ids:
                        if attachment.mimetype and "image" in attachment.mimetype:
                            # try:
                            # Decode the base64 image data
                            image_data = base64.b64decode(attachment.datas)

                            # Resize the image
                            max_width = 100  # Maximum width in pixels
                            max_height = 100  # Maximum height in pixels

                            # Open the image using PIL
                            image = PILImage.open(io.BytesIO(image_data))
                            width, height = image.size
                            aspect_ratio = width / height

                            # Resize while maintaining aspect ratio
                            if width > max_width:
                                width = max_width
                                height = int(width / aspect_ratio)

                            if height > max_height:
                                height = max_height
                                width = int(height * aspect_ratio)

                            resized_image = image.resize((width, height), PILImage.Resampling.LANCZOS)

                            # Convert resized image to BytesIO for OpenPyXL
                            img_bytes = io.BytesIO()
                            resized_image.save(img_bytes, format='PNG')
                            img_bytes.seek(0)

                            # Add the image to the worksheet
                            cell_location = f'I{cur_row}'
                            img = Image(img_bytes)
                            ws.add_image(img, cell_location)


                            # Adjust row height to fit the image
                            ws.row_dimensions[cur_row].height = 80  # Adjust as needed

                        # except Exception as e:
                        #     _logger.error(f"Error processing image {attachment.name}: {e}")



                # Apply styling
                for col in 'ABCDEFGHIJ':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')



                cur_row += 1
        except Exception as e:
            _logger.error("Error filling verification list: %s", str(e))

    # ----- Strong Points -----
    def _create_strong_points(self, ws):
        columns = {
            'A': ('Clause', 20),
            'B': ('Checkpoints', 40),
            'C': ('Observations/Findings', 30)
        }
        self._create_sheet_structure(ws, columns)
        self._fill_strong_points(ws)
        ws.merge_cells('C5:D5')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')

    def _fill_strong_points(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for point in self.strong_points_ids:
                ws.merge_cells(f'C{cur_row}:D{cur_row}')
                ws[f'A{cur_row}'] = point.clause if point.clause else ''
                ws[f'B{cur_row}'] = point.checkpoints if point.checkpoints else ''
                ws[f'C{cur_row}'] = point.observations if point.observations else ''

                for col in 'ABC':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cur_row += 1
        except Exception as e:
            _logger.error("Error filling strong points: %s", str(e))

    # ----- Nonconformities -----
    def _create_nonconformities_sheet(self, ws):
        columns = {
            'A': ('Reference', 30),
            'B': ('Name', 20),
            'C': ('Created on', 20),
            'D': ('Filled in by', 30),
            'E': ('Responsible', 20),
            'F': ('Manager', 20),
            'G': ('System', 20),
            'H': ('Stage', 20)
        }
        self._create_sheet_structure(ws, columns)
        self._fill_nonconformities_sheet(ws)

    def _fill_nonconformities_sheet(self, ws):
        styles = self._get_common_styles()
        cur_row = 6

        try:
            for nc in self.nonconformity_ids:
                ws[f'A{cur_row}'] = nc.ref or ''
                ws[f'B{cur_row}'] = nc.name or ''
                ws[f'C{cur_row}'] = nc.create_date.strftime('%Y-%m-%d') if nc.create_date else ''
                ws[f'D{cur_row}'] = nc.user_id.name if nc.user_id else ''
                ws[f'E{cur_row}'] = nc.responsible_user_id.name if nc.responsible_user_id else ''
                ws[f'F{cur_row}'] = nc.manager_user_id.name if nc.manager_user_id else ''
                ws[f'G{cur_row}'] = nc.system_id.name if nc.system_id else ''
                ws[f'H{cur_row}'] = nc.stage_id.name if nc.stage_id else ''

                for col in 'ABCDEFGH':
                    cell = ws[f'{col}{cur_row}']
                    cell.border = styles['border']
                    cell.alignment = styles['align_center']
                    cell.alignment = Alignment(wrap_text=True)
                cur_row += 1
        except Exception as e:
            _logger.error("Error filling nonconformities: %s", str(e))

    # ----- Image Processing -----
    # def _process_company_logo(self, image_data, max_size=(120, 120), padding=(10, 10)):
    #
    #     try:
    #         img = PILImage.open(BytesIO(image_data))
    #
    #         # Resize image while maintaining aspect ratio
    #         img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
    #
    #         # Add padding to align the image properly
    #         img_with_padding = ImageOps.expand(img, border=padding, fill='white')
    #
    #         # Save the processed image to a BytesIO object
    #         temp_bytes = BytesIO()
    #         img_with_padding.save(temp_bytes, format='PNG')
    #         temp_bytes.seek(0)
    #
    #         # Return the processed image as openpyxl's Image object
    #         return Image(temp_bytes)
    #     except Exception as e:
    #         _logger.error("Logo processing failed: %s", str(e))
    #         return None
    def get_resized_img(self,image, cell, ws, max_width=400, max_height=65, padding_top=10, padding_left=10):
        """
        Retrieve and resize the company logo while maintaining aspect ratio.
        Adds padding on the top and left side.

        :param max_width: Maximum allowed width of the logo
        :param max_height: Maximum allowed height of the logo
        :param padding_top: Extra padding at the top
        :param padding_left: Extra padding at the left
        :return: openpyxl Image object (logo) or None if no logo is found
        """
        # if not self.env.user.company_id.logo:
        #     return None  # No logo available

        image_data = base64.b64decode(image)
        image = PILImage.open(io.BytesIO(image_data))

        # Get original dimensions and aspect ratio
        width, height = image.size
        aspect_ratio = width / height

        # Resize while maintaining aspect ratio
        if width > max_width:
            width = max_width
            height = int(width / aspect_ratio)

        if height > max_height:
            height = max_height
            width = int(height * aspect_ratio)

        # Resize image
        resized_image = image.resize((width, height), PILImage.LANCZOS)

        # Add padding (transparent background)
        padded_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill=(255, 255, 255, 0))

        # Save to bytes
        img_bytes = io.BytesIO()
        padded_image.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        logo_image = Image(img_bytes)
        ws.add_image(logo_image, cell)


    def add_logo_to_worksheet(self, ws, cell="E1"):
        """
        Adds the resized company logo to the specified Excel cell.

        :param ws: openpyxl worksheet object
        :param cell: Cell location where the logo should be placed (default: 'A1')
        """
        logo_image = self.get_resized_logo()
        if logo_image:
            ws.add_image(logo_image, cell)



    def _process_image(self, image_data, max_size=(100, 100)):
        try:
            img = PILImage.open(BytesIO(image_data))
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            temp_bytes = BytesIO()
            img.save(temp_bytes, format='PNG')
            return Image(temp_bytes)
        except Exception as e:
            _logger.error("Image processing failed: %s", str(e))
            return None
