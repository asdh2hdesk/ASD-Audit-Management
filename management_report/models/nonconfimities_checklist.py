import base64
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from bs4 import BeautifulSoup
from PIL import ImageOps
from PIL import Image as PILImage
import io

from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.styles.builtins import styles
from openpyxl.worksheet.datavalidation import DataValidation
from odoo import api, fields, models
from bs4 import BeautifulSoup



_logger = logging.getLogger(__name__)


class NonconformitiesReport(models.Model):
    _inherit = 'mgmtsystem.nonconformity'

    generate_xls_file = fields.Binary(string="Generated File")



    def action_generate_nonconformities_report(self):
        """
        Main function to generate the Excel report.
        """
        output = BytesIO()
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Create sheets
        sheets = {
            'Description': self._create_description_sheet,
            'Procedures': self._create_procedures_sheet,
            'Related Audits': self._create_related_audits_sheet,
        }

        for sheet_name, create_func in sheets.items():
            try:
                ws = wb.create_sheet(sheet_name)
                create_func(ws)
                self._style_and_fill_sheet(ws)
                self.get_resized_img(self.env.user.company_id.logo, 'E3', ws)
            except Exception as e:
                _logger.error(f"Error creating sheet {sheet_name}: {e}")

        # Save the workbook to binary content
        wb.save(output)
        output.seek(0)
        self.generate_xls_file = base64.b64encode(output.getvalue()).decode('utf-8')

        # Download the Excel file
        return {
            'type': 'ir.actions.act_url',
            'target': 'self',
            'url': f'/web/content?model=mgmtsystem.nonconformity&id={self.id}&filename=Nonconformities_Report.xlsx&field=generate_xls_file&download=true'
        }

    def _get_common_styles(self):
        """
        Get reusable styles for formatting the Excel sheets.
        """
        return {
            'header_fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
            'table_header_fill': PatternFill(start_color="003366", end_color="003366", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'header_font': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'normal_font': Font(name='Arial', size=10),
            'align_left': Alignment(horizontal='left', vertical='center'),
            'align_center': Alignment(horizontal='center', vertical='center')
        }

    def _style_and_fill_sheet(self, ws):
        """
        Apply styles and add generic content to a worksheet.
        """
        styles = self._get_common_styles()

        # Set column widths
        for col in 'ABCDE':
            ws.column_dimensions[col].width = 25

        # Add the title
        ws.merge_cells('A1:G2')
        ws['A1'] = "Nonconformities Report"
        ws['A1'].font = styles['header_font']
        ws['A1'].fill = styles['header_fill']
        ws['A1'].alignment = styles['align_center']

        # Add headers with default styles
        headers = [
            ('A3', 'Name:', 'B3', getattr(self, 'name', 'N/A') or 'N/A'),
            ('A4', 'Reference:', 'B4', getattr(self, 'ref', 'N/A') or 'N/A'),
            ('A5', 'Created on:', 'B5', self.create_date.strftime('%Y-%m-%d') if self.create_date else 'N/A'),
            ('A6', 'Partner:', 'B6', self.partner_id.name if self.partner_id else 'N/A'),
            ('A7', 'Related to:', 'B7', self.reference if self.reference else 'N/A'),
            ('A8', 'Origin :', 'B8', ', '.join(self.origin_ids.mapped('name')) if self.origin_ids else 'N/A'),
            ('C3', 'Responsible:', 'D3', self.responsible_user_id.name if self.responsible_user_id else 'N/A'),
            ('C4', 'Manager:', 'D4', self.manager_user_id.name if self.manager_user_id else 'N/A'),
            ('C5', 'Filled in by:', 'D5', self.user_id.name if self.user_id else 'N/A'),
            ('C6', 'System:', 'D6', self.system_id.name if self.system_id else 'N/A'),
            ('C7', 'Company:', 'D7', self.company_id.name if self.company_id else 'N/A'),
            ('C8', 'Status:', 'D8',
             {'normal': 'In Progress', 'done': 'Ready for next stage', 'blocked': 'Blocked'}.get(self.kanban_state,
                                                                                                 'N/A'))
        ]

        for label_ref, label, value_ref, value in headers:
            ws[label_ref] = label
            ws[label_ref].font = styles['header_font']
            ws[label_ref].fill = styles['header_fill']
            ws[label_ref].border = styles['border']
            ws[label_ref].alignment = styles['align_center']

            ws[value_ref] = value
            ws[value_ref].font = styles['normal_font']
            ws[value_ref].border = styles['border']
            ws[value_ref].alignment = styles['align_center']



    def _create_sheet_structure(self, ws, columns):
        """
        Create the basic structure of a sheet with column headers and widths.
        """
        styles = self._get_common_styles()


        for col, (header, width) in columns.items():
            ws.column_dimensions[col].width = width
            cell = ws.cell(row=9, column=ord(col) - ord('A') + 1, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['table_header_fill']
            cell.border = styles['border']
            cell.alignment = styles['align_center']


    def parse_html_to_text_list(self, html_content):
        """
        Parse HTML content and return a list of text elements.

        :param html_content: str, HTML content to parse
        :return: list of str, extracted text elements
        """
        # Create a BeautifulSoup object
        soup = BeautifulSoup(html_content, 'html.parser')

        # Extract text from the soup object
        text_elements = soup.find_all(text=True)

        # Filter out unwanted elements (like script, style, etc.)
        filtered_text = [text.strip() for text in text_elements if
                         text.strip() and text.parent.name not in ['script', 'style']]

        return filtered_text

    def _create_description_sheet(self, ws):
        """
        Creates the Description sheet with a formatted header and structured data.
        """
        # Define styles for the header
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')


        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Merge cells for the "Description" header
        ws.merge_cells('A9:G9')
        ws['A9'] = "Description"
        ws['A9'].font = header_font
        ws['A9'].fill = header_fill
        ws['A9'].alignment = header_alignment
        ws['A9'].border = header_border

        # Add column headers
        headers = ["Line", "Content"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Fill description data
        self._fill_description_data(ws)

    def _fill_description_data(self, ws):
        """
        Extract and format Description data from HTML and add it in a structured format.
        """
        styles = self._get_common_styles()
        row = 10  # Start adding data from row 3 (after headers)

        # Check if description exists
        if self.description:
            # Parse HTML content into a list of text elements
            description_parsed = self.parse_html_to_text_list(self.description)

            if description_parsed:
                # Split the parsed content into key-value pairs
                description_dict = {}
                current_key = None

                for line in description_parsed:
                    if ":" in line:  # Check if the line contains a key-value pair
                        key, value = line.split(":", 1)  # Split only on the first occurrence of ":"
                        description_dict[key.strip()] = value.strip()
                        current_key = key.strip()
                    elif current_key:  # If the line is a continuation of the previous value
                        description_dict[current_key] += " " + line.strip()

                # Insert parsed description data in multiple rows
                for key, value in description_dict.items():
                    # Add key in column A
                    ws[f'A{row}'] = key
                    ws[f'A{row}'].alignment = styles['align_left']
                    ws[f'A{row}'].font = styles['normal_font']
                    ws[f'A{row}'].border = styles['border']

                    # Add value in column B
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].alignment = styles['align_left']
                    ws[f'B{row}'].font = styles['normal_font']
                    ws[f'B{row}'].border = styles['border']
                    ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    ws.merge_cells(f'B{row}:G{row}')
                    row += 1
            else:
                # If no parsed content, add a placeholder
                ws[f'A{row}'] = "No description available"
                ws[f'A{row}'].alignment = styles['align_left']
                ws[f'A{row}'].font = styles['normal_font']
                ws[f'A{row}'].border = styles['border']

                ws.merge_cells(f'B{row}:G{row}')
                ws[f'B{row}'] = "No description available"
                ws[f'B{row}'].alignment = styles['align_left']
                ws[f'B{row}'].font = styles['normal_font']
                ws[f'B{row}'].border = styles['border']
        else:
            # If no description, add a placeholder
            ws[f'A{row}'] = "No description available"
            ws[f'A{row}'].alignment = styles['align_left']
            ws[f'A{row}'].font = styles['normal_font']
            ws[f'A{row}'].border = styles['border']

            ws.merge_cells(f'B{row}:G{row}')
            ws[f'B{row}'] = "No description available"
            ws[f'B{row}'].alignment = styles['align_left']
            ws[f'B{row}'].font = styles['normal_font']
            ws[f'B{row}'].border = styles['border']





    def _create_procedures_sheet(self, ws):
        """
        Add content to the Procedures sheet.
        """

        columns = {
            'A': ('Title', 30),
            'B': ('Category', 20),
            'C': ('Last Updated By', 30),
            'D': ('Last Updated On', 20),
            'E': ('Company', 35),
        }
        self._create_sheet_structure(ws, columns)

        self._fill_procedures_data(ws)
        ws.merge_cells('E9:G9')

        ws['E9'].alignment = Alignment(horizontal='center', vertical='center')

    def _fill_procedures_data(self, ws):
        """
        Populate the Procedures sheet with data.
        """
        styles = self._get_common_styles()
        row = 10  # Start after the column headers

        for procedure in self.procedure_ids:
            ws.merge_cells(f'E{row}:G{row}')

            ws[f'A{row}'] = procedure.name or ''
            ws[f'B{row}'] = procedure.parent_id.name or ''
            ws[f'C{row}'] = procedure.write_uid.name if procedure.write_uid else ''
            ws[f'D{row}'] = procedure.write_date.strftime('%Y-%m-%d') if procedure.write_date else ''
            ws[f'E{row}'] = procedure.company_id.name if procedure.company_id else ''

            for col in 'ABCDEFG':
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                cell.alignment = styles['align_center']
            row += 1

    def _create_related_audits_sheet(self, ws):
        """
        Add content to the Related Audits sheet.


        """

        columns = {
            'A': ('Reference', 30),
            'B': ('Name', 30),
            'C': ('Planned Date', 20),
            'D': ('System', 20),
            'E': ('Company', 20),
            'F': ('City', 20),
            'G': ('Plant Code', 20),
            'H': ('Status', 20),
        }
        self._create_sheet_structure(ws, columns)
        self._fill_related_audits_data(ws)


    def _fill_related_audits_data(self, ws):

       styles = self._get_common_styles()
       row= 10
       for audit in self.audit_ids:
           ws[f'A{row}'] = audit.reference or ''
           ws[f'B{row}'] = audit.name or ''
           ws[f'C{row}'] = audit.date.strftime('%Y-%m-%d') if audit.date else ''
           ws[f'D{row}'] = audit.system_id.name if audit.system_id else ''
           ws[f'E{row}'] = audit.company_id.name if audit.company_id else ''
           ws[f'F{row}'] = audit.company_state or ''
           ws[f'G{row}'] = audit.plant_name or ''
           ws[f'H{row}'] = audit.state or ''
           for col in 'ABCDEFGH':
               cell = ws[f'{col}{row}']
               cell.border = styles['border']
               cell.alignment = styles['align_center']
           row += 1

    def get_resized_img(self, image, cell, ws, max_width=400, max_height=80, padding_top=10, padding_left=10):
        """
        Retrieve and resize the company logo while maintaining aspect ratio.
        Adds padding on the top and left side.

        :param max_width: Maximum allowed width of the logo
        :param max_height: Maximum allowed height of the logo
        :param padding_top: Extra padding at the top
        :param padding_left: Extra padding at the left
        :return: openpyxl Image object (logo) or None if no logo is found
        """
        # if not self.env.user.company_id.logo:
        #     return None  # No logo available

        image_data = base64.b64decode(image)
        image = PILImage.open(io.BytesIO(image_data))

        # Get original dimensions and aspect ratio
        width, height = image.size
        aspect_ratio = width / height

        # Resize while maintaining aspect ratio
        if width > max_width:
            width = max_width
            height = int(width / aspect_ratio)

        if height > max_height:
            height = max_height
            width = int(height * aspect_ratio)

        # Resize image
        resized_image = image.resize((width, height), PILImage.LANCZOS)

        # Add padding (transparent background)
        padded_image = ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill=(255, 255, 255, 0))

        # Save to bytes
        img_bytes = io.BytesIO()
        padded_image.save(img_bytes, format='PNG')
        img_bytes.seek(0)
        logo_image = Image(img_bytes)
        ws.add_image(logo_image, cell)

    def add_logo_to_worksheet(self, ws, cell="D1"):
        """
        Adds the resized company logo to the specified Excel cell.

        :param ws: openpyxl worksheet object
        :param cell: Cell location where the logo should be placed (default: 'A1')
        """
        logo_image = self.get_resized_logo()
        if logo_image:
            ws.add_image(logo_image, cell)

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            # Apply the border to the cells where the logo is placed
            ws[cell].border = border


    def _process_image(self, image_data, max_size=(100, 100)):
        try:
            img = PILImage.open(BytesIO(image_data))
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            temp_bytes = BytesIO()
            img.save(temp_bytes, format='PNG')
            return Image(temp_bytes)
        except Exception as e:
            _logger.error("Image processing failed: %s", str(e))
            return None




